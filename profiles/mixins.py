from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import View
from myvmeste.settings import DEBUG
from .models import Profile, Orders, ApplicationsForMoney, Cards
from .forms import ApplicationsForMoneyForm, UploadFileForm, UploadOrderPayForm
from web.utils import automatic_report
from openpyxl import load_workbook


class ProfileMixin(LoginRequiredMixin, View):
    login_url = '/profile/login'

    def __init__(self):
        super(ProfileMixin, self).__init__()
        self.application_form = ApplicationsForMoneyForm()
        self.order_form = UploadFileForm()
        self.order_form_pay = UploadOrderPayForm()
        self.user = None
        self.context = {
            'application_form': self.application_form,
            'order_form': self.order_form,
            'order_form_pay': self.order_form_pay,
        }

    def post(self, request, *args, **kwargs):
        if 'report' in request.FILES:
            self.download_report(request)
        if 'report_pay' in request.FILES:
            self.download_report_pay(request)
        if 'reserved' in request.POST:
            application_form = ApplicationsForMoneyForm(request.POST)
            card = request.POST['card']
            user = request.user
            if application_form.is_valid():
                reserved = application_form.cleaned_data['reserved']
            else:
                reserved = None
            balance = user.profile.available()
            if reserved and balance >= reserved:
                user.profile.balance.available -= reserved
                if user.profile.balance.available < 0:
                    difference = user.profile.balance.available
                    user.profile.balance.self_available += difference
                    user.profile.balance.available = 0
                user.profile.balance.reserved += reserved
                ApplicationsForMoney.objects.create(user=user, card=card, reserved=reserved)
                user.profile.balance.save()
                messages.add_message(
                    request,
                    messages.INFO,
                    '{} руб зарезервированы для вывода и будут выплачены в течении 7 дней'.format(reserved))
            else:
                messages.add_message(request, messages.ERROR, 'Ошибка вывода денежных средств')

    def get_context_data(self, request):
        self.user = request.user
        cards = Cards.objects.filter(user=self.user).select_related('user')
        self.context['cards'] = cards
        self.context['user'] = self.user

    def download_report(self, request):
        """ Загрузка и парсинг отчета партнерки в формате excel
        """
        item_row, item_user, money = 0, 0, 0
        self.order_form = UploadFileForm(request.POST, request.FILES)
        new_upload_order = Orders()
        if self.order_form.is_valid():
            new_upload_order.order = request.FILES['report']
            new_upload_order.save()
            path = './media/uploads/{}' if DEBUG else '/home/www/site/myvmeste_info/media/uploads/{}'
            ex_data = load_workbook(path.format(request.FILES['report']))
            first_sheet = ex_data.get_sheet_names()[0]
            worksheet = ex_data.get_sheet_by_name(first_sheet)

            for row in worksheet.iter_rows():
                is_personal_sale, order_id, status, offer_id = False, None, None, None
                item_row += 1
                if item_row == 1:
                    continue
                item_col = 0
                for cell in row:
                    item_col += 1
                    if item_col == 1:
                        order_id = cell.value
                    elif item_col == 3:
                        offer_id = int(cell.value.split(',')[0])
                    elif item_col == 6:
                        item_user = self.get_user_from_table_data(cell)
                    elif item_col == 7:
                        sub2_value = self.get_user_from_table_data(cell)
                        if sub2_value:
                            item_user = sub2_value
                            is_personal_sale = True
                    elif item_col == 11:
                        status = cell.value
                automatic_report(order_id, item_user.user, status, offer_id, is_personal_sale)
            new_upload_order.delete()

    def get_user_from_table_data(self, cell):
        try:
            item_user = Profile.objects.get(pk=int(cell.value))
        except TypeError:
            return None
        return item_user

    def download_report_pay(self, request):
        """ Загрузка отчета выплаченных средств
        """
        item_row = 0
        item_application = None
        item_balance = None
        item_user = None
        new_upload_order = Orders()
        self.order_form_pay = UploadOrderPayForm(request.POST, request.FILES)

        if self.order_form_pay.is_valid():
            new_upload_order.order = request.FILES['report_pay']
            new_upload_order.save()
            path = './media/uploads/{}' if DEBUG else '/home/www/site/myvmeste_info/media/uploads/{}'
            ex_data = load_workbook(path.format(request.FILES['report_pay']))
            first_sheet = ex_data.get_sheet_names()[0]
            worksheet = ex_data.get_sheet_by_name(first_sheet)

            for row in worksheet.iter_rows():
                item_row += 1
                if item_row == 1 or item_row == 2:
                    continue
                item_col = 0
                for cell in row:
                    item_col += 1
                    if item_col == 1:
                        item_application = ApplicationsForMoney.objects.get(pk=cell.value)
                        item_user = item_application.user.profile
                        item_balance = item_user.balance
                    elif item_col == 7:
                        status = cell.value
                        if status == 'Да' or status == 'да' or status == '1':
                            money = item_application.reserved
                            item_balance.accrued -= money
                            item_balance.paid_out += money
                            item_balance.save()
                            item_application.is_paid_out = True
                            item_application.save()
            new_upload_order.delete()
